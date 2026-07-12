import pandas as pd
import matplotlib.pyplot as plt
plt.style.use('dark_background')
import glob
import os
import sys

    

POWER_COLOR = "4682B4"
VOLTAGE_COLOR = "FFA500"
POWER_MATPLOTLIB_COLOR = "steelblue"
VOLTAGE_MATPLOTLIB_COLOR = "orange"


if len(sys.argv) > 1:
    FILE = sys.argv[1]
else:
    files = glob.glob("wind_data_*.csv")
    FILE = max(files, key=os.path.getctime)
print(f"Loading: {FILE}") 
file_base = os.path.splitext(os.path.basename(FILE))[0]
chart_file = f"{file_base}_chart.png"
excel_file = f"{file_base}.xlsx"

df = pd.read_csv(FILE) # File CSV table is now df
i = 0
df["Time in S"] = [(i) * 0.2 for i in range(len(df))] # adds column Time in S
time = df["Time in S"] # df = from the table
Power = df["Power_mW"]
Voltage = df["RealVoltage_V"]

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8,4)) # sublots = 2 graphs in 1 img, ax1 and ax2 being the graphs

ax1.plot(                            # First Graph Design
    time,                             # X variable
    Power,                            # Y variable
    color=POWER_MATPLOTLIB_COLOR,     # graph color
    linewidth=1.5, 
    marker="o", 
    markersize=5,
    label="Power")
for x,y in zip(time,Power):            
    if y > 0:
        ax1.annotate(f"{y}", (x, y))    # shows variable stats

ax1.set_title(                      # title adjustments
    "Power Over Time",
    color = "black",
    size = 14,
    weight = "bold")

ax1.set_xlabel(                    #label adjusments
    "Time (s)",
    color = "Green",
    size = 13,
    weight = "bold")

ax1.set_ylabel( 
    "Power (mW)",
    color = "Red",
    size = 13,
    weight = "bold")

ax1.minorticks_on()           #activate small grids
ax1.grid(                      #grid adjustments
    True, 
    which = "both" ,
    linestyle="-", 
    color = "black", 
    alpha=0.5,
    linewidth = 0.5)
ax1.legend()

ax2.plot(
    time, 
    Voltage, 
    color=VOLTAGE_MATPLOTLIB_COLOR, 
    linewidth=1.5, 
    marker="o", 
    markersize=5,
    label="Voltage")

for x,y in zip(time, Voltage):
    if y > 0:
        ax2.annotate(f"{y}", (x,y))

ax2.set_title(
    "Voltage Over Time",
    color = "black",
    size = 13,
    weight = "bold")
ax2.set_xlabel(
    "Time (s)",
    color = "Green",
    size = 11,
    weight = "bold")
ax2.set_ylabel(
    "Voltage (V)",
    color = "blue",
    size = 11,
    weight = "bold")

ax2.minorticks_on()
ax2.grid(
    True, 
    which = "both",
    linestyle="-", 
    color = "black", 
    alpha=0.5,
    linewidth = 0.5)
ax2.legend()



plt.tight_layout()               #layout cleanup
plt.savefig(chart_file, dpi = 400) # saves the Matplotlib graph image
print(f"Saved chart image: {chart_file}")

try:
    import openpyxl # allow editing and creating excel files
    from openpyxl.chart import LineChart, Reference # create editable Excel graphs
    from openpyxl.utils.dataframe import dataframe_to_rows # transfer data to excel

    def style_chart_series(chart, color):
        series = chart.series[0]
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 20000
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color

    workbook = openpyxl.Workbook() # creates brand new excel file
    ws_data = workbook.active #grab first sheet
    ws_data.title = 'Data'
    for rows in dataframe_to_rows(df,index=False, header = True): #insert df in workbook, for loop to go row by row
        ws_data.append(rows)

    ws_chart = workbook.create_sheet("Graph")
    max_row = ws_data.max_row
    time_data = Reference(ws_data, min_col=6, min_row=2, max_row=max_row)

    power_chart = LineChart()
    power_chart.title = "Power Over Time"
    power_chart.y_axis.title = "Power (mW)"
    power_chart.x_axis.title = "Time (s)"
    power_values = Reference(ws_data, min_col=5, min_row=1, max_row=max_row)
    power_chart.add_data(power_values, titles_from_data=True)
    power_chart.set_categories(time_data)
    style_chart_series(power_chart, POWER_COLOR)
    power_chart.legend = None
    power_chart.width = 14
    power_chart.height = 7
    ws_chart.add_chart(power_chart, "B2")

    voltage_chart = LineChart()
    voltage_chart.title = "Voltage Over Time"
    voltage_chart.y_axis.title = "Voltage (V)"
    voltage_chart.x_axis.title = "Time (s)"
    voltage_values = Reference(ws_data, min_col=3, min_row=1, max_row=max_row)
    voltage_chart.add_data(voltage_values, titles_from_data=True)
    voltage_chart.set_categories(time_data)
    style_chart_series(voltage_chart, VOLTAGE_COLOR)
    voltage_chart.legend = None
    voltage_chart.width = 14
    voltage_chart.height = 7
    ws_chart.add_chart(voltage_chart, "B18")

    workbook.save(excel_file) # save to disk as xlsx
    print(f"Saved Excel file: {excel_file}")
except ModuleNotFoundError:
    print(f"Excel export skipped: install openpyxl to create {excel_file}.")

plt.subplots_adjust(hspace=0.4)
plt.show()
